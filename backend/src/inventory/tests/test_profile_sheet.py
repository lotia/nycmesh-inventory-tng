"""Tests for the workbook reader and the report it feeds.

The real workbook carries volunteer names and email addresses, is gitignored,
and is not ours to publish, so nothing here reads it. Every test builds its
own workbook holding exactly the rows the property under test needs, which is
also what makes the population rule visible: a fixture with one furniture row
in it says what the rule does in a way a count of 3,439 cannot.
"""

import io
from pathlib import Path
from typing import Any

import pytest
from django.core.management import call_command
from django.core.management.base import CommandError
from openpyxl import Workbook

from inventory.sheet import workbook
from inventory.tests import sheets
from inventory.tests.sheets import AT

HEADERS = (
    "Timestamp",
    "Email Address",
    "First name",
    "Checking in/Checking out",
    "Item",
    "How many?",
    "Planned Use / Project / Notes",
)


def build(tmp_path: Path, submissions: list[tuple[Any, ...]], catalogue: tuple[str, ...] = ("LiteBeam",)) -> Path:
    """Write a workbook shaped like the export, holding just these rows."""
    book = Workbook()
    items = book.active
    assert items is not None
    items.title = workbook.CATALOGUE_TAB
    # Column D, with the QR link in C, because that pairing is the trap the
    # reader exists to get right and a fixture holding only D would not catch
    # a reader that read C.
    items.append(("Add new rows", "streakwave", "QR", "Name"))
    for name in catalogue:
        items.append(("", "streakwave", "docs.google.com/forms/...", name))

    responses = book.create_sheet(workbook.SUBMISSIONS_TAB)
    responses.append(HEADERS)
    for row in submissions:
        responses.append(row)

    path = tmp_path / "sheet.xlsx"
    book.save(path)
    return path


def submission(direction: str = workbook.CHECKING_OUT, **fields: Any) -> tuple[Any, ...]:
    """One row of the tab, in the column order the reader is being tested on.

    Built over `sheets.submission` rather than beside it, so that there is one
    set of defaults and one `AT`. The column order is the part this module
    genuinely owns: it is what the reader is being asserted against, and a
    reader that read the columns in a different order would have to disagree
    with this tuple rather than with a builder it shares.
    """
    row = sheets.submission(direction=direction, **fields)
    return (row.at, row.email, row.name, row.direction, row.item, row.quantity, row.note)


def test_a_row_carrying_no_direction_is_not_a_submission(tmp_path: Path) -> None:
    """The seventeen rows the tab holds beyond the population are furniture:
    a heading somebody typed, a testing form, a warning against typing rows by
    hand. Counting them is how a figure stops agreeing with decision 0008.
    """
    path = build(
        tmp_path,
        [
            submission(),
            ("ADD NEW PRODUCTS HERE TO PREVENT ERRORS--->", "", "", "", "LiteBeam", "", ""),
            submission(workbook.CHECKING_IN),
        ],
    )

    sheet = workbook.read(path)

    assert sheet.rows_read == 3
    assert len(sheet.submissions) == 2
    assert sheet.without_direction == 1
    assert (sheet.check_outs, sheet.check_ins) == (1, 1)


def test_a_fully_blank_row_is_not_read_at_all(tmp_path: Path) -> None:
    """Blank rows are what a spreadsheet leaves behind, not something anybody
    entered, so they are outside the count the population rule divides.
    """
    path = build(tmp_path, [submission(), (None, None, None, None, None, None, None)])

    sheet = workbook.read(path)

    assert sheet.rows_read == 1


def test_the_catalogue_is_read_from_the_name_column_and_not_the_qr_link(tmp_path: Path) -> None:
    """A reader taking the wrong column still returns 52 of something, which
    is why ``workbook``'s header names the two apart.
    """
    path = build(tmp_path, [submission()], catalogue=("LiteBeam", "OmniTikPOE"))

    sheet = workbook.read(path)

    assert sheet.catalogue == ("LiteBeam", "OmniTikPOE")


def test_a_row_stopping_short_of_the_note_is_still_read(tmp_path: Path) -> None:
    """openpyxl hands back only the cells the sheet filled, so a submission
    with nothing in the notes field arrives as a shorter tuple.
    """
    path = build(tmp_path, [(AT, "a@example.net", "Ada", workbook.CHECKING_OUT, "LiteBeam", 1.0)])

    sheet = workbook.read(path)

    assert sheet.submissions[0].note == ""


def test_every_field_arrives_trimmed_and_never_none(tmp_path: Path) -> None:
    """Normalising in the reader is what stops six rules each having their own
    answer to whether ' Ada ' and 'Ada' are one person.
    """
    path = build(tmp_path, [submission(email="  a@example.net ", name=" Ada ", note="  mesh room  ", quantity=None)])

    read = workbook.read(path).submissions[0]

    assert (read.email, read.name, read.note) == ("a@example.net", "Ada", "mesh room")
    assert read.quantity is None


def test_whitespace_inside_a_cell_is_collapsed_too(tmp_path: Path) -> None:
    """``workbook._text`` names the character this uses and what it defeats."""
    path = build(tmp_path, [(AT, "", "", "Checking\xa0Out", "TP-Link\xa0SFP-RJ45", 1.0, "mesh\xa0room")])

    sheet = workbook.read(path)

    assert len(sheet.submissions) == 1
    assert sheet.submissions[0].item == "TP-Link SFP-RJ45"
    assert sheet.submissions[0].note == "mesh room"


def test_a_true_in_the_quantity_column_is_not_a_quantity_of_one(tmp_path: Path) -> None:
    """A bool is an int in Python, so the obvious isinstance check reads TRUE
    as one of something rather than as the nothing it is.
    """
    path = build(tmp_path, [submission(quantity=True)])

    assert workbook.read(path).submissions[0].quantity is None


def test_a_file_that_is_not_a_workbook_says_so(tmp_path: Path) -> None:
    """The brief tells contributors their export may arrive as .ods or .csv,
    and openpyxl's own message for one of those names a zip archive.
    """
    csv = tmp_path / "sheet.csv"
    csv.write_text("Timestamp,Email Address\n")

    with pytest.raises(CommandError, match=r"not a readable \.xlsx workbook"):
        call_command("profile_sheet", str(csv))


def test_a_workbook_missing_a_tab_names_the_tab(tmp_path: Path) -> None:
    """A renamed tab otherwise escapes as a KeyError traceback, which is the
    failure the path guard beside it exists to prevent.
    """
    book = Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.title = "Something Else"
    path = tmp_path / "sheet.xlsx"
    book.save(path)

    with pytest.raises(CommandError, match=f"no {workbook.CATALOGUE_TAB} and no {workbook.SUBMISSIONS_TAB} tab"):
        call_command("profile_sheet", str(path))


def test_a_submission_knows_the_spreadsheet_row_it_came_from(tmp_path: Path) -> None:
    """A figure checked by hand is checked against the row number the sheet
    shows, so the reader counts from the first row below the header.
    """
    path = build(tmp_path, [submission(), submission()])

    assert [s.row for s in workbook.read(path).submissions] == [2, 3]


def test_a_timestamp_that_is_not_one_is_read_as_absent(tmp_path: Path) -> None:
    """A hand-typed row can carry anything in the timestamp column, and the
    batching rule sorts on it.
    """
    path = build(tmp_path, [submission(at="not a date")])

    assert workbook.read(path).submissions[0].at is None


def test_the_report_states_the_population_it_counted(tmp_path: Path) -> None:
    path = build(
        tmp_path,
        [submission(), submission(workbook.CHECKING_IN), ("", "", "", "", "LiteBeam", "", "testing form")],
    )
    out = io.StringIO()

    call_command("profile_sheet", str(path), stdout=out)

    # Read back with the padding collapsed: the alignment is there for a
    # reader and pinning it here would make widening a label a test failure.
    printed = [" ".join(line.split()) for line in out.getvalue().splitlines()]
    assert "Population" in printed
    assert "rows on QRresponses 3" in printed
    assert "carrying a direction 2" in printed
    assert "carrying neither 1" in printed
    assert "catalogued items 1" in printed


def test_a_path_that_is_not_a_workbook_is_refused_by_name(tmp_path: Path) -> None:
    """openpyxl's own message for a missing file names a zip archive, which
    reads as a corrupt workbook rather than a path typed wrong.
    """
    with pytest.raises(CommandError, match="No workbook at"):
        call_command("profile_sheet", str(tmp_path / "nothing.xlsx"))
