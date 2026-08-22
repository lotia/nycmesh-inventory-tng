"""The workbook reader, and the population rule every figure rests on.

Two tabs matter. `Fast Inventory` holds the catalogue, and the item name is
**column D**; the column beside it holds the QR link and has the same number of
filled rows, which
[the brief](../../../../docs/briefs/sheet-classifiers.md) records has already
caught a reader. `QRresponses` holds the form submissions, one row per item
taken or returned.

The population rule is the one
[decision 0008](../../../../docs/decisions/0008-stock-ledger-transfer-graph.md#context)
states, and the counts that rest on it are stated there too rather than
repeated here: a row counts when it carries a direction. The rest of the tab
is sheet furniture -- `ADD NEW PRODUCTS HERE TO PREVENT ERRORS--->`, a
`testing form` row, a warning against typing rows by hand -- and is not
submissions anybody made.
"""

from collections.abc import Iterator, Mapping
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from inventory.sheet import Report

CATALOGUE_TAB = "Fast Inventory"
SUBMISSIONS_TAB = "QRresponses"

# Column D of the catalogue tab, zero-based, per the docstring above.
CATALOGUE_NAME_COLUMN = 3

CHECKING_OUT = "Checking Out"
CHECKING_IN = "Checking In"
DIRECTIONS = (CHECKING_OUT, CHECKING_IN)


@dataclass(frozen=True)
class Submission:
    """One row of `QRresponses`, read into the seven columns the form wrote.

    Read for every non-blank row, including the furniture: what the direction
    column of a rejected row actually said is the only thing that answers why
    it was rejected. `SubmissionRow.taken` applies the population rule and
    `Sheet.submissions` is what it leaves.

    Two of the seven arrive as the sheet stored them rather than as the model
    wants them, because converting is a decision to make and to record rather
    than one to make in a reader:

    - `quantity` is a float, because the form asked for a number and Sheets
      kept it as one. A movement needs an integer.
    - `at` is **naive**, because every timestamp in the export is, and this
      project runs with `USE_TZ` on. Which zone they were written in is not in
      the workbook; assuming one here would bury the assumption in a reader
      nobody reads. It is also `None` where a row was typed by hand, so a rule
      that sorts on it has to say what it does with those.
    """

    row: int
    at: datetime | None
    email: str
    name: str
    direction: str
    item: str
    quantity: float | None
    note: str

    @property
    def is_check_out(self) -> bool:
        return self.direction == CHECKING_OUT


@dataclass(frozen=True)
class CatalogueRow:
    """One non-blank row of `Fast Inventory`, and the name read out of it.

    `name` is empty where the row fills other columns and not the name column.
    Such a row is part of the tab and is not a catalogued item, so it is read
    and `Sheet.catalogue` leaves it out.
    """

    number: int
    cells: tuple[Any, ...]
    name: str


@dataclass(frozen=True)
class SubmissionRow:
    """One non-blank row of `QRresponses`: what it holds, and what was read.

    `cells` is the row as the sheet stores it, kept beside the reading because
    the reading is lossy by design -- `_text` collapses, the quantity column
    is narrowed to a number, and a direction that is not one is dropped from
    `Sheet.submissions` entirely. A staging table built on the reading alone
    could not answer what the row actually said.
    """

    cells: tuple[Any, ...]
    read: Submission

    @property
    def taken(self) -> bool:
        """Whether the population rule counts this row."""
        return self.read.direction in DIRECTIONS


@dataclass(frozen=True)
class Sheet:
    """A workbook read into the two things the rules ask questions about.

    Held as rows and read down to those two, rather than the other way about,
    so that whatever reconstructs a sheet from somewhere other than a workbook
    -- `inventory.staging` -- is handing over the same thing the reader does.
    """

    catalogue_rows: tuple[CatalogueRow, ...]
    rows: tuple[SubmissionRow, ...]

    @property
    def catalogue(self) -> tuple[str, ...]:
        return tuple(row.name for row in self.catalogue_rows if row.name)

    @property
    def submissions(self) -> tuple[Submission, ...]:
        return tuple(row.read for row in self.rows if row.taken)

    @property
    def rows_read(self) -> int:
        """Non-blank rows below the header, whether or not they carry a direction.

        Reported so that the population rule is shown doing something rather
        than asserted as a number nobody can check.
        """
        return len(self.rows)

    @property
    def without_direction(self) -> int:
        return self.rows_read - len(self.submissions)

    @property
    def check_outs(self) -> int:
        return sum(1 for s in self.submissions if s.is_check_out)

    @property
    def check_ins(self) -> int:
        return len(self.submissions) - self.check_outs


def by_column(cells: tuple[Any, ...]) -> dict[str, Any]:
    """A row's cells, keyed by the column letters the spreadsheet shows.

    Here rather than with the caller that stores them, because a column letter
    is a fact about a spreadsheet and this module is the only one that knows
    it is reading one. Letters rather than positions, and rather than the
    header row's own names, so that a cell being checked by hand is checked
    as `D812` -- what the person has in front of them -- and so that a tab
    whose headers are reworded still names the same keys.
    """
    return {get_column_letter(index): cell for index, cell in enumerate(cells, start=1)}


def in_column_order(keyed: Mapping[str, Any]) -> tuple[Any, ...]:
    """`by_column` undone.

    Sorted on the letter rather than trusting the order the keys arrive in: a
    caller that stored them has no promise of getting them back in the order
    it wrote them, and that short keys tend to come first is not one.
    """
    return tuple(cell for _, cell in sorted(keyed.items(), key=lambda pair: column_index_from_string(pair[0])))


def _text(value: Any) -> str:
    """The cell as the rule will see it: a string, whitespace collapsed, never None.

    Normalising here rather than in each rule is deliberate. Every rule would
    otherwise have to remember to, and the one that forgot would report a
    figure a point or two away from the others for a reason nobody could see.

    Collapsed rather than merely trimmed, because the whitespace that breaks a
    rule is the whitespace in the middle. A non-breaking space is what a value
    pasted from a web page carries, and `Checking\xa0Out` is not a direction,
    `TP-Link\xa0SFP-RJ45` reaches no alias, and `mesh\xa0room` is nowhere.
    Today's export is clean; that is not a property of the next one.
    """
    return "" if value is None else " ".join(str(value).split())


def _cell(row: tuple[Any, ...], index: int) -> Any:
    """Column `index` of `row`, or None where the row stops short.

    openpyxl gives a row only as many cells as the sheet filled, so a
    submission with an empty note is a six-tuple rather than a seven-tuple
    holding None.
    """
    return row[index] if index < len(row) else None


def _blank(row: tuple[Any, ...]) -> bool:
    return all(_text(cell) == "" for cell in row)


def _catalogue(rows: Iterator[tuple[Any, ...]]) -> list[CatalogueRow]:
    return [
        CatalogueRow(number=number, cells=row, name=_text(_cell(row, CATALOGUE_NAME_COLUMN)))
        # Enumerated from 2 so that a row number here is the row number the
        # spreadsheet shows, which is what somebody checking a figure by hand
        # has in front of them. Both tabs count the same way.
        for number, row in enumerate(rows, start=2)
        if not _blank(row)
    ]


def _rows(rows: Iterator[tuple[Any, ...]]) -> list[SubmissionRow]:
    read = []
    for number, row in enumerate(rows, start=2):
        if _blank(row):
            continue
        at = _cell(row, 0)
        quantity = _cell(row, 5)
        read.append(
            SubmissionRow(
                cells=row,
                read=Submission(
                    row=number,
                    at=at if isinstance(at, datetime) else None,
                    email=_text(_cell(row, 1)),
                    name=_text(_cell(row, 2)),
                    direction=_text(_cell(row, 3)),
                    item=_text(_cell(row, 4)),
                    # `not isinstance(quantity, bool)` because a bool is an int
                    # in Python, so a TRUE in the column would arrive as one of
                    # something rather than as the nothing it is.
                    quantity=float(quantity)
                    if isinstance(quantity, int | float) and not isinstance(quantity, bool)
                    else None,
                    note=_text(_cell(row, 6)),
                ),
            ),
        )
    return read


class NotTheWorkbook(Exception):
    """The file at that path is not the export this reads.

    Raised rather than let openpyxl's own message out, for the same reason
    profile_sheet checks the path exists: `BadZipFile: File is not a zip file`
    reads as a corrupt workbook to somebody who has in fact handed over a
    `.csv`, which the brief tells contributors their export may arrive as.
    """


def read(path: Path) -> Sheet:
    """Read the workbook at `path`.

    `read_only` because the workbook is fifteen megabytes of QR images nothing
    here looks at, and `data_only` because the catalogue's stock column is a
    formula: without it openpyxl hands back `=SUMIF(...)` rather than the
    number the sheet shows.
    """
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as opening:
        raise NotTheWorkbook(f"{path} is not a readable .xlsx workbook: {opening}") from opening
    try:
        missing = [tab for tab in (CATALOGUE_TAB, SUBMISSIONS_TAB) if tab not in workbook.sheetnames]
        if missing:
            raise NotTheWorkbook(f"{path} has no {' and no '.join(missing)} tab.")
        # read_only trusts the <dimension> record the producing application
        # wrote, and a workbook whose record understates the used range comes
        # back short with no error at all. Today's export has none, so resetting
        # costs a scan of the two tabs and buys a silent truncation.
        catalogue_tab, submissions_tab = workbook[CATALOGUE_TAB], workbook[SUBMISSIONS_TAB]
        catalogue_tab.reset_dimensions()
        submissions_tab.reset_dimensions()
        catalogue = _catalogue(catalogue_tab.iter_rows(min_row=2, values_only=True))
        rows = _rows(submissions_tab.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()
    return Sheet(catalogue_rows=tuple(catalogue), rows=tuple(rows))


def section(sheet: Sheet) -> Report:
    """Which rows count, before any rule has an opinion about what they say.

    Here rather than in the command, so that every rule's report is a section
    of its own module and the importer can have the population without
    importing a management command.
    """
    return "Population", [
        (f"rows on {SUBMISSIONS_TAB}", sheet.rows_read),
        ("  carrying a direction", len(sheet.submissions)),
        (f"    {CHECKING_OUT}", sheet.check_outs),
        (f"    {CHECKING_IN}", sheet.check_ins),
        ("  carrying neither", sheet.without_direction),
        ("catalogued items", len(sheet.catalogue)),
    ]
